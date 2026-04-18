import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment
from ingest import load_orders

OUTPUT_PATH = '../output/order_report.xlsx'

def generate_report(df):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Order Summary"

    # --- Header Row ---
    headers = ['Order ID', 'Customer', 'Product', 'Quantity', 'Due Date', 'Status', 'Missing Data', 'Duplicate']
    header_fill = PatternFill("solid", fgColor="2C3E50")
    header_font = Font(color="FFFFFF", bold=True)

    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center')

    # --- Data Rows ---
    red_fill = PatternFill("solid", fgColor="FF9999")
    green_fill = PatternFill("solid", fgColor="99FF99")

    for row_idx, (_, order) in enumerate(df.iterrows(), start=2):
        values = [
            order['order_id'],
            order['customer'],
            order['product'],
            order['quantity'],
            order['due_date'],
            order['status'],
            'Yes' if order['missing_data'] else 'No',
            'Yes' if order['is_duplicate'] else 'No'
        ]

        is_problem = order['missing_data'] or order['is_duplicate']

        for col_idx, value in enumerate(values, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.fill = red_fill if is_problem else green_fill
            cell.alignment = Alignment(horizontal='center')

    # --- Column Widths ---
    for col in ws.columns:
        max_length = max(len(str(cell.value or '')) for cell in col)
        ws.column_dimensions[col[0].column_letter].width = max_length + 4

    # --- Summary Section ---
    total = len(df)
    flagged = len(df[df['missing_data'] | df['is_duplicate']])
    valid = total - flagged

    ws.append([])
    ws.append(['SUMMARY', '', '', '', '', '', '', ''])
    ws.append(['Total Orders', total])
    ws.append(['Valid Orders', valid])
    ws.append(['Flagged Orders', flagged])

    wb.save(OUTPUT_PATH)
    print(f"✅ Report saved to {OUTPUT_PATH}")

# Run it
df = load_orders('../data/orders.csv')
generate_report(df)